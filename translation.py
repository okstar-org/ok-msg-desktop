#  Copyright (c) 2022 船山信息 chuanshaninfo.com
#  The project is licensed under Mulan PubL v2.
#  You can use this software according to the terms and conditions of the Mulan
#  PubL v2. You may obtain a copy of Mulan PubL v2 at:
#           http://license.coscl.org.cn/MulanPubL-2.0
#  THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
#  EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
#  MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
#  See the Mulan PubL v2 for more details.

import html
import openpyxl
import os
import re
import shutil
import subprocess
import xml.etree.ElementTree as ET
from enum import Enum
from pathlib import Path
from xml.dom import minidom


class TranslateTitle(Enum):
    ColumnBegin = 0
    Translate_Code = 1
    Translate_Chinese = 2
    Translate_English = 3
    Translate_ChineseTw = 4
    Translate_French = 5
    Translate_Korean = 6
    Translate_Japanese = 7
    ColumnEnd = 8

    Title = 1    # 第一行为标题
    RowBegin = 2


class TranslationManager:

    def __init__(self):
        self.trans_info_hash = {}

    def get_trans_info_hash(self, excel_file_path):
        self.trans_info_hash.clear()

        # 加载 Excel 文件
        try:
            xlsx = openpyxl.load_workbook(excel_file_path)
        except Exception as e:
            print(f"ERROR: fail to load xlsx doc: {e}")
            return False

        sheet_names = xlsx.sheetnames
        if not sheet_names:
            print("ERROR: no sheets found in xlsx file")
            return False

        # 获取第一个工作表
        current_sheet = xlsx[sheet_names[0]]
        if current_sheet is None:
            print("currentSheet is NULL")
            return False

        # 获取行列数
        row_count = current_sheet.max_row
        column_count = current_sheet.max_column

        if column_count < TranslateTitle.ColumnBegin.value or row_count < TranslateTitle.RowBegin.value:
            print(f"{row_count} {column_count}")
            return False

        # 处理每一行
        for i in range(TranslateTitle.RowBegin.value, row_count + 1):
            key = current_sheet.cell(row=i, column=TranslateTitle.Translate_Code.value).value
            if not key:
                continue

            dest_hash = {}

            for title in range(TranslateTitle.Translate_Code.value + 1, TranslateTitle.ColumnEnd.value):
                dest = current_sheet.cell(row=i, column=title).value
                dest_hash[title] = dest

            if key in self.trans_info_hash and self.trans_info_hash[key] == dest_hash:
                # DEBUG: 如果内容一致则跳过
                pass
            else:
                self.trans_info_hash[key] = dest_hash

        return True

    def update_ts_file_by_excel(self, ts_file_path, language):
        # 加载 .ts 文件
        ts_file = Path(ts_file_path)
        if not ts_file.exists():
            print(f"Failed to open ts file: {ts_file_path}")
            return

        try:
            tree = ET.parse(ts_file)
            root = tree.getroot()
        except Exception as e:
            print(f"Failed to parse .ts file: {e}")
            return

        # 设置语言属性
        root.set("language", ts_file.stem)

        # 更新翻译内容
        for context in root.findall('context'):
            for message in context.findall("message"):
                # 查找 source 元素
                source_elem = message.find("source")
                # 查找 translation 元素
                translation_elem = message.find("translation")
                
                # 获取 source_text，默认为空字符串
                source_text = source_elem.text if source_elem is not None and source_elem.text else ""
                
                # 如果 source_text 存在并且在 trans_info_hash 中
                if source_text and source_text in self.trans_info_hash:
                    new_text = self.trans_info_hash[source_text].get(language, "")
                    
                    # 如果没有新翻译，跳过
                    if not new_text:
                        continue

                    # 获取 translation 元素的 type 属性
                    if translation_elem is not None:
                        type_attr = translation_elem.get("type")
                        
                        # 如果翻译类型是 unfinished，移除 type 属性
                        if type_attr == "unfinished":
                            translation_elem.attrib.pop("type", None)

                    # 查找 numElem 元素
                    num_elem = translation_elem.find("numerusform") if translation_elem is not None else None

                    if num_elem is None:  # 如果没有 numElem 元素
                        if translation_elem is not None:
                            # 如果翻译元素的文本为空或只有空白字符，则更新其文本
                            if not translation_elem.text or not translation_elem.text:
                                translation_elem.text = new_text
                            else:
                                translation_elem.text = new_text
                    else:
                        num_elem.text = new_text

        xml_str = ET.tostring(root, 'utf-8')
        parsed_xml = minidom.parseString(xml_str)  # 使用 minidom 解析
        pretty_xml = parsed_xml.toprettyxml(indent="    ")  # 设置缩进为 4 个空格

        # 去除多余的空行
        pretty_xml = re.sub(r'\n\s*\n', '\n', pretty_xml)
        # 使用 html.unescape 去除转义字符
        pretty_xml = html.unescape(pretty_xml)

        # 保存到文件
        with open(ts_file_path, "w", encoding="utf-8") as out_file:
            out_file.write(pretty_xml)

        # 插入注释
        comment = '''<?xml version='1.0' encoding='utf-8'?>
<!--
  ~ Copyright (c) 2022 船山信息 chuanshaninfo.com
  ~ The project is licensed under Mulan PubL v2.
  ~ You can use this software according to the terms and conditions of the Mulan
  ~ PubL v2. You may obtain a copy of Mulan PubL v2 at:
  ~          http://license.coscl.org.cn/MulanPubL-2.0
  ~ THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
  ~ EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
  ~ MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
  ~ See the Mulan PubL v2 for more details.
  -->

<!DOCTYPE TS>'''
        
        self.replace_first_line_in_file(ts_file_path, comment)

        print(f"Translations updated successfully: {ts_file_path}")

    def replace_first_line_in_file(self, file_path, comment):
        """ 替换文件的第一行 """
        
        # 读取文件内容
        with open(file_path, "r", encoding="utf-8") as file:
            content = file.readlines()

        # 替换第一行的文本
        content[0] = comment + "\n"  # 替换第一行

        # 将修改后的内容写回文件
        with open(file_path, "w", encoding="utf-8") as file:
            file.writelines(content)


    def process_files(self, src_path):
        def func(ts_path):
            # 调用 get_trans_info_hash 和 update_ts_file_by_excel
            self.get_trans_info_hash(ts_path + "translation.xlsx")
            self.update_ts_file_by_excel(ts_path + "ja.xml", TranslateTitle.Translate_Japanese.value)
            self.update_ts_file_by_excel(ts_path + "zh_TW.xml", TranslateTitle.Translate_ChineseTw.value)
            self.update_ts_file_by_excel(ts_path + "ko.xml", TranslateTitle.Translate_Korean.value)
            self.update_ts_file_by_excel(ts_path + "zh_CN.xml", TranslateTitle.Translate_Chinese.value)

        # 定义路径并调用 func
        ts_paths = [
            "UI/window/login/ts/",
            "UI/window/main/ts/",
            "UI/window/config/ts/",
            "modules/platform/ts/",
            "modules/im/ts/",
            "modules/meet/ts/"
        ]

        for path in ts_paths:
            func(src_path + path)


def run_lupdate(src_path, ts_files):
    # 定义lupdate命令，逐个添加路径和目标文件
    command = ["lupdate"]
    command.append(src_path)  # 添加多个源路径
    command.append("-ts")  # 添加-ts选项
    command.extend(ts_files)  # 添加多个翻译文件
    command.extend(["-no-obsolete", "-locations", "none"])  # 添加其他选项

    # 执行命令并获取输出
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        
        # 打印标准输出
        print("STDOUT:", result.stdout)
        
        # 打印标准错误（如果有）
        if result.stderr:
            print("STDERR:", result.stderr)

        # 处理TS文件：删除同名XML文件，并将TS文件重命名为XML文件
        process_ts_files(ts_files)

    except subprocess.CalledProcessError as e:
        print(f"Command failed with error: {e}")
        print("STDERR:", e.stderr)
    except Exception as e:
        print(f"An error occurred: {e}")


def process_ts_files(ts_files):
    for ts_file in ts_files:
        xml_file = os.path.splitext(ts_file)[0] + ".xml"
        # 如果同名的XML文件存在，则删除它
        if os.path.exists(xml_file):
            os.remove(xml_file)
            print(f"Removed existing XML file: {xml_file}")
        
        # 将TS文件重命名为XML文件
        if os.path.exists(ts_file):
            shutil.move(ts_file, xml_file)
            print(f"Renamed TS file to XML: {xml_file}")


# 主程序入口
if __name__ == "__main__":
    
    # 定义源路径和目标翻译文件列表
    def update_ts_files(src_path, languages=["zh_CN", "zh_TW", "ko", "ja"]):
        ts_files = [f"{src_path}ts/{lang}.ts" for lang in languages]
        run_lupdate(src_path, ts_files)


    update_ts_files("src/UI/login/")
    update_ts_files("src/UI/main/")
    update_ts_files("src/modules/config/")
    update_ts_files("./src/modules/platform/")
    update_ts_files("./src/modules/im/")
    update_ts_files("./src/modules/meet/")

    src_path = "./src/"
    translation_manager = TranslationManager()
    translation_manager.process_files(src_path)
